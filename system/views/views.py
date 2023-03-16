from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from system.models.person import Person
from system.models.event import Event
from system.models.register import Register
from system.models.company import Company
from system.models.check_in import CheckIn
from system.forms import CompanyForm, PersonForm, QuestionForm, CnpjForm, CpfForm


@login_required
def home(request):
    return render(request, 'home/welcome.html')


def credential_maker(request, id_event, id_person):
    person = get_object_or_404(Person, id=id_person)
    event = Event.objects.get(id=id_event)
    return render(request, 'credential/credential.html', {'person': person, 'event':event })


@login_required
def report_generator(request, id_event):
    event = Event.objects.get(id=id_event)
    registers = Register.objects.filter(event=event)
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=Report_UMV.xlsx'

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Relatório do evento " + event.title

    columns = [
        'Nome',
        'E-mail',
        'Telefone',
        'Cargo',
        'Empresa',
        'CNPJ',
        'Check-in',
        'Inscrição'
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for register in registers:
        row_num += 1
        row = [
            register.person.name,
            register.person.email,
            register.person.phone,
            register.person.role,
            register.person.company.fantasy_name,
            register.person.company.cnpj,
            register.checkin,
            register.status
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response            

def r_register(request, id_event, id_company):
    company = Company.objects.get(id=id_company)
    event = Event.objects.get(id=id_event)
    form = CpfForm(request.POST or None)
    if form.is_valid():
        try:
            person = Person.objects.get(cpf=form.cleaned_data['cpf'])          
        except:
            person = None
        if person:
            registered = Register.objects.filter(
                person=person, event=event).count()
            current_stocking = Register.objects.filter(event=event).count()
            if  current_stocking < event.qtd_registration:
                if registered == 0:
                    Register.objects.create(
                        person=person,
                        event=event
                    )
                    return render(request, 'register/register_success.html')
                else:
                    return render(request, 'register/register_waiting_approval.html', {'event': event})
            else:
                return render(request, 'register/max_stocking.html', {'event': event})
        else:
            return redirect('r_person_create', id_event=event.id)
    return render(request, 'register/r_register.html', {'form': form, 'event': event, 'company': company})

def r_person_create(request, id_company, id_event):
    event = Event.objects.get(id=id_event)
    form = PersonForm(request.POST or None)
    company = Company.objects.get(id=id_company)
    if form.is_valid():
        person = form.save(commit=False)
        person.company = company
        person.save()
        Register.objects.create(
            person=person,
            event=event
        )
        return render(request, 'register/r_person_create_success.html',{'event':event})
    data = {
        'form': form,
        'company': company,
        'event': event
    }
    return render(request, 'register/r_person_create.html', data)

def checkin(request, id):
    event = Event.objects.get(id=id)
    form = CpfForm(request.POST or None)
    if form.is_valid():
        try:
            person = Person.objects.get(cpf=form.cleaned_data['cpf'])
        except:
            person = None
        if person:
            try:
                registered = Register.objects.get(
                    event=event, person=person)
            except:
                registered = None
            if registered:
                if not registered.approved:
                    return render(request, 'register/register_waiting_approval.html', {'event': event})
                else:    
                    if registered.checked_in:
                        return render(request, 'checkin/already_checkin.html', {'event': event})                                            
                CheckIn.objects.create(
                    person=person,
                    event=event,
                    register=registered
                )
                return render(request, 'checkin/checkin_success.html', {'event': event})                  
            else:
                return render(request, 'checkin/checkin_fail.html', {'event': event})
        else:
            return redirect('r_company_search', event.id)
    return render(request, 'checkin/checkin.html', {'form': form, 'event': event})

def r_person_create(request, id_event):
    event = Event.objects.get(id=id_event)
    form = PersonForm(request.POST or None)
    company = Company.objects.get(id=1)
    if form.is_valid():
        person = form.save(commit=False)
        person.company = company
        person.save()
        current_stocking = Register.objects.filter(event=event).count()
        if  current_stocking < event.qtd_registration:
            Register.objects.create(
                person=person,
                event=event
            )
            # return render(request, 'register/r_person_create_success.html',{'event':event})
            return render(request, 'register/register_success.html')
        else:
            return render(request, 'register/max_stocking.html', {'event': event})
    data = {
        'form': form,
        'company': company,
        'event': event
    }
    return render(request, 'register/r_person_create.html', data)