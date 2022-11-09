from django.shortcuts import render, redirect, get_object_or_404
from .models import (Question, Company, Person, Event, Register,
                     RafflePrizes, CheckIn)
from .forms import (QuestionForm, CompanyForm, PersonForm, EventForm,
                    CpfForm)
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib import messages
from system.filters import CompanyFilter
import random
from openpyxl import Workbook
from weasyprint import HTML
from django.http import HttpResponse

@login_required
def home(request):
    return render(request, 'home/welcome.html')

# Registro
@login_required
def register_approve(request, id_register):
    register = Register.objects.get(id=id_register)
    company = register.person.company
    event = register.event
    person = register.person
    if company.approved:
        if register.approved:
            register.approved = False
            register.save()
            messages.success(request, 'Status de aprovação alterado, o registro foi reprovado.')
            return redirect('event_detail', id_event=event.id)
        else:
            register.approved = True
            register.save()
            msg_html = render_to_string('mail/approved_person.html', {
                'person': person,
                'company': company,
                'event': event
            })
            send_mail(
                'Aprovação de registro',
                msg_html,
                'umv@appdoevento.com.br',
                [person.email],
                fail_silently=False,
                html_message=msg_html,
            )
            messages.success(request, 'Status de aprovação alterado, o registro foi aprovado.')        
            return redirect('event_detail', id_event=event.id)
    else:
        messages.success(request, 'Empresa do participante ainda não foi aprovada. Aprove a empresa para continuar com o processo.')              
        return redirect('event_detail', id_event=event.id)

@login_required
def register_delete(request, id_register):
    register = Register.objects.get(id=id_register)
    event = register.event
    register.delete()
    messages.success(request, 'Registro excluído com sucesso!')
    return redirect(event_detail, id_event=event.id)



def company_search(request, id_event):
    event = Event.objects.get(id=id_event)
    cfilter = CompanyFilter(
        request.GET, queryset=Company.objects.filter(approved=True))
    num_results = Company.objects.filter(approved=True).count()
    data = {
        'cfilter': cfilter,
        'num_results': num_results,
        'event': event
    }
    return render(request, 'register/company_search.html', data)

@login_required
def register(request, id_event, id_company):  
    company = Company.objects.get(id=id_company)
    event = Event.objects.get(id=id_event)
    form = CpfForm(request.POST or None)
    if form.is_valid():
        try:
            person = Person.objects.get(cpf=form.cleaned_data['cpf'])
        except:
            person = None
        if person:
            registered = Register.objects.filter(
                person=person, event=event).count()
            current_stocking = Register.objects.filter(event=event).count()
            if  current_stocking < event.qtd_registration:
                if registered == 0:
                    Register.objects.create(
                        person=person,
                        event=event
                    )
                    messages.success(request, 'Registro efetuado com sucesso!')
                    return redirect('event_detail', id_event=event.id)
                else:
                    messages.success(request, 'Usuário já resgistrado neste evento!')
                    return redirect('event_detail', id_event=event.id)
            else:
                messages.success(request, 'Evento atingiu a lotação máxima, remova participantes ou aumente a capadidade do evento')
                return redirect('event_detail', id_event=event.id)
        else:
            return redirect('person_create_and_register', id_company=company.id, id_event=event.id)
    return render(request, 'register/register.html', {'form': form, 'event': event, 'company': company})


# Question Views
@login_required
def question_list(request):
    events = Event.objects.all()
    return render(request, 'questions/questions.html', {'events': events})

@login_required
def question_by_event(request, id_event):
    event = Event.objects.get(id=id_event)
    questions = Question.objects.filter(event=event)    
    return render(request, 'questions/question_list.html', {'questions': questions, 'event': event})  

@login_required
def question_approve(request, id_event, id_question):
    event = Event.objects.get(id=id_event)
    question = Question.objects.get(id=id_question)
    if question.approve == True:
        question.approve = False    
    else:
        question.approve = True        
    question.save()
    return redirect('question_by_event', id_event=id_event)

@login_required
def question_delete(request, id_event, id_question):
    event = Event.objects.get(id=id_event)
    question = Question.objects.get(id=id_question)
    question.delete()
    return redirect('question_by_event', id_event=id_event)


@login_required
def question_approved(request):
    events = Event.objects.all()
    return render(request, 'questions/question_approved.html', {'events': events})    

@login_required
def question_approved_by_event(request, id_event):
    events = Event.objects.get(id=id_event)
    question = Question.objects.filter(event=events, approve=True)
    return render(request, 'questions/question_approved_by_event.html', {'questions': question, 'events': events})

@login_required
def question_clean(request, id_event):
    event = Event.objects.get(id=id_event)
    question = Question.objects.filter(event=event)
    question.delete()
    return redirect('question_by_event', id_event=id_event)

# Company
@login_required
def company_create(request):
    form = CompanyForm(request.POST or None)
    if form.is_valid():
        company = form.save(commit=False)
        form.save()
        return redirect('company_detail', company.id)
    data = {
        'form': form
    }
    return render(request, 'company/company_create.html', data)

@login_required
def company_approve(request, id):
    company = Company.objects.get(id=id)
    company.approved = True
    company.save()
    return redirect('company_list')

@login_required
def company_list(request):
    cfilter = CompanyFilter(
        request.GET, queryset=Company.objects.all())
    data = {
        'cfilter': cfilter
    }
    return render(request, 'company/company_list.html', data)

@login_required
def company_detail(request, id):
    company = Company.objects.get(id=id)
    persons = Person.objects.filter(company=company)
    data = {
        'company': company,
        'persons': persons
    }
    return render(request, 'company/company_detail.html', data)

@login_required
def company_update(request, id):
    company = Company.objects.get(id=id)
    form = CompanyForm(request.POST or None, instance=company)
    if form.is_valid():
        form.save()
        return redirect('company_list')
    data = {
        'form': form,
        'company': company
    }
    return render(request, 'company/company_update.html', data)

@login_required
def company_delete(request, id):
    company = Company.objects.get(id=id)
    company.delete()
    return redirect('company_list')

# Person
@login_required
def person_list(request):
    persons = Person.objects.all()
    data = {
        'persons': persons
    }
    return render(request, 'person/person_list.html', data)

@login_required
def person_create(request, id_company):
    company = Company.objects.get(id=id_company)
    form = PersonForm(request.POST or None)
    if form.is_valid():
        person = form.save(commit=False)
        person.company = company
        person.save()       
        return redirect('company_detail', company.id)
    data = {
        'form': form,
        'company': company
    }
    return render(request, 'person/person_create.html', data)


@login_required
def person_create_and_register(request, id_company, id_event):
    event = Event.objects.get(id=id_event)
    company = Company.objects.get(id=id_company)    
    form = PersonForm(request.POST or None)
    if form.is_valid():
        person = form.save(commit=False)
        person.company = company
        person.save()
        Register.objects.create(
            person=person,
            event=event,
            approved=True
        )
        messages.success(request, 'Participante incluído com sucesso')        
        return redirect('event_detail', id_event=id_event)
    data = {
        'form': form,
        'company': company,
        'event': event
    }
    return render(request, 'person/person_create_and_register.html', data)        

@login_required
def person_update(request, id):
    person = Person.objects.get(id=id)
    id_company = person.company.id
    form = PersonForm(request.POST or None, instance=person)
    if form.is_valid():
        form.save()
        return redirect('company_detail', id=id_company)
    data = {
        'form': form,
        'person': person
    }
    return render(request, 'person/person_update.html', data)

@login_required
def person_delete(request, id):
    company = Person.objects.get(id=id).company.id
    person = Person.objects.get(id=id)
    person.delete()
    return redirect('company_detail', id=company)

# Event
@login_required
def event_list(request):
    events = Event.objects.all()
    data = {
        'events': events
    }
    return render(request, 'event/event_list.html', data)

@login_required
def event_create(request):
    form = EventForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('event_list')
    data = {
        'form': form
    }
    return render(request, 'event/event_create.html', data)

@login_required
def event_update(request, id):
    event = Event.objects.get(id=id)
    form = EventForm(request.POST or None, instance=event)
    if form.is_valid():
        form.save()
        return redirect('event_list')
    data = {
        'form': form,
        'event': event
    }
    return render(request, 'event/event_update.html', data)

@login_required
def event_detail(request, id_event):
    event = Event.objects.get(id=id_event)
    registers = Register.objects.filter(event=event)
    data = {
        'event': event,
        'registers': registers
    }
    return render(request, 'event/event_detail.html', data)

@login_required
def event_delete(request, id):
    event = Event.objects.get(id=id)
    event.delete()
    return redirect('event_list')

#Sorteio
@login_required
def raffle_event_list(request):
    events = Event.objects.all()
    return render(request, 'raffle/raffle.html', {'events': events})

@login_required
def raffle_prepare(request, id_event):
    event = Event.objects.get(id=id_event)
    checkeds_in = CheckIn.objects.filter(event=event)
    winners = RafflePrizes.objects.filter(event=event)
    return render(request, 'raffle/make_raffle.html', {'event': event, 'checkeds_in':checkeds_in, 'winners':winners })

@login_required
def raffle_maker(request, id_event):
    event = Event.objects.get(id=id_event)
    checkeds_in = CheckIn.objects.filter(event=event)
    participants = []

    for people in checkeds_in:
        participants.append(people.person.id)
    id_winner = random.choice(participants)
    winner = Person.objects.get(id=id_winner)

    already_won = RafflePrizes.objects.filter(person=winner).count()
    qtd_winners = RafflePrizes.objects.filter(event=event).count()
    qtd_participants = CheckIn.objects.filter(event=event).count()
    if qtd_participants > qtd_winners:
        if already_won == 0:
            RafflePrizes.objects.create(
                event=event,
                person=winner
            )
        else:
            raffle_maker(request, event.id)        
    return redirect('raffle_prepare', id_event=id_event)


@login_required
def raffle_delete(request, id_event, id_winner):
    event = Event.objects.get(id=id_event)
    person = Person.objects.get(id=id_winner)
    delete = RafflePrizes.objects.filter(event=event, person=person)
    delete.delete()
    return redirect('raffle_prepare', id_event=id_event)


def credential_maker(request, id_event, id_person):
    person = get_object_or_404(Person, id=id_person)
    event = Event.objects.get(id=id_event)
    return render(request, 'credential/credential.html', {'person': person, 'event':event })


@login_required
def report_generator(request, id_event):
    event = Event.objects.get(id=id_event)
    registers = Register.objects.filter(event=event)
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=Report_UMV.xlsx'

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Relatório do evento " + event.title

    columns = [
        'Nome',
        'E-mail',
        'Telefone',
        'Cargo',
        'Empresa',
        'CNPJ',
        'Check-in',
        'Inscrição'
    ]

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for register in registers:
        row_num += 1
        row = [
            register.person.name,
            register.person.email,
            register.person.phone,
            register.person.role,
            register.person.company.fantasy_name,
            register.person.company.cnpj,
            register.checkin,
            register.status
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response            

